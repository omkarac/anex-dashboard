"""
Skygauge — output writers.

Three pluggable backends, same interface:
  * ExcelWriter    — multi-sheet .xlsx file. Best for local validation runs.
  * CsvWriter      — single CSV. Simplest, universal, no openpyxl needed.
  * SupabaseWriter — production target. Spatial indexes + RPCs + Realtime.

The scraper picks one via the --output flag:
    --output excel:./data/skygauge.xlsx
    --output csv:./data/skygauge.csv
    --output supabase

All writers are idempotent on noc_id (re-running the scraper will not duplicate
records). Excel and CSV both save after each (airport × structure × window) so
that a crashed run preserves partial progress.
"""

from __future__ import annotations

import csv
import logging
import os
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Optional, Protocol

if TYPE_CHECKING:
    from nocas_scraper import ParsedNOC, ScrapeWindow

logger = logging.getLogger("skygauge.writers")

# ---------------------------------------------------------------------------
# Column orders — keep these stable so existing Excel files keep working
# ---------------------------------------------------------------------------

NOC_COLUMNS: list[str] = [
    "noc_id", "sacfa_id", "airport_code",
    "lat", "lon",
    "site_elevation_m", "permissible_top_m", "permissible_top_raw",
    "is_restricted", "structure_type", "status",
    "issue_date", "issue_date_known",
    "pdf_url", "owner_name", "site_address",
    "scraped_at",
]

LOG_COLUMNS: list[str] = [
    "run_id", "tracking_id", "scraper", "trigger_source",
    "airport_code", "structure_type", "from_date", "to_date",
    "started_at", "completed_at", "duration_seconds",
    "records_added", "records_updated", "errors", "status", "notes",
]

AIRPORT_COLUMNS: list[str] = [
    "aai_id", "code", "name", "arp_lat", "arp_lon", "elevation_m",
]


# AAI airport ID → human-readable code mapping. Lives here so writers can
# render the readable code in the airport_code column without a DB round-trip.
AAI_ID_TO_CODE: dict[int, str] = {
    15:  "VABB",   # Santa Cruz (CSMIA)
    16:  "VAJJ",   # Juhu
    140: "VANM",   # Navi Mumbai
}


# ---------------------------------------------------------------------------
# Writer protocol
# ---------------------------------------------------------------------------

class OutputWriter(Protocol):
    """All writers expose the same surface area for the scraper orchestrator."""

    def get_airport_id(self, aai_id: int) -> Any:
        """Return whatever identifier this backend uses to reference an airport.
        For Excel/CSV that's the AAI ID itself; for Supabase it's the serial PK."""
        ...

    def log_run_start(
        self,
        scraper_name: str,
        window: "ScrapeWindow",
        airport_id: Any,
        tracking_id: Optional[str],
        trigger_source: str,
    ) -> str:
        """Insert a log row; return its run_id."""
        ...

    def log_run_complete(
        self,
        run_id: str,
        added: int,
        updated: int,
        errors: int,
        duration_ms: int,
        status: str,
        notes: str = "",
    ) -> None:
        ...

    def upsert_nocs(self, airport_id: Any, parsed: list["ParsedNOC"]) -> tuple[int, int]:
        """Returns (added, updated_or_skipped)."""
        ...

    def close(self) -> None:
        """Flush and release any resources (e.g. final Excel save)."""
        ...


# ---------------------------------------------------------------------------
# Shared row builders
# ---------------------------------------------------------------------------

def _now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def noc_to_row(p: "ParsedNOC", airport_code: str) -> dict:
    """Convert a ParsedNOC into a flat dict matching NOC_COLUMNS order."""
    return {
        "noc_id":              p.noc_id,
        "sacfa_id":            p.sacfa_id or "",
        "airport_code":        airport_code,
        "lat":                 p.lat,
        "lon":                 p.lon,
        "site_elevation_m":    p.site_elevation_m if p.site_elevation_m is not None else "",
        "permissible_top_m":   p.permissible_top_m if p.permissible_top_m is not None else "",
        "permissible_top_raw": p.permissible_top_raw or "",
        "is_restricted":       bool(p.is_restricted),
        "structure_type":      p.structure_type,
        "status":              p.status,
        "issue_date":          p.issue_date.isoformat() if p.issue_date else "",
        "issue_date_known":    p.issue_date is not None,
        "pdf_url":             p.pdf_url or "",
        "owner_name":          p.owner_name or "",
        "site_address":        p.site_address or "",
        "scraped_at":          _now_iso(),
    }


# ===========================================================================
# Excel writer (recommended for local backfill validation runs)
# ===========================================================================

class ExcelWriter:
    """Multi-sheet .xlsx output. Atomic-save pattern: write to .tmp then rename."""

    def __init__(self, path: Path) -> None:
        try:
            import openpyxl  # noqa: F401  (only needed for excel mode)
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is required for --output excel:... "
                "Install with: pip install openpyxl"
            ) from e
        self.path = Path(path).expanduser().resolve()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._load_or_create()
        self._existing_noc_ids: Optional[set[str]] = None

    def _load_or_create(self) -> None:
        import openpyxl
        if self.path.exists():
            logger.info("Loading existing workbook: %s", self.path)
            self.wb = openpyxl.load_workbook(self.path)
            for sheet in ("nocs", "scrape_log", "airports"):
                if sheet not in self.wb.sheetnames:
                    self._create_sheet(sheet)
            return

        logger.info("Creating new workbook: %s", self.path)
        self.wb = openpyxl.Workbook()
        # The default sheet becomes "nocs"
        ws = self.wb.active
        ws.title = "nocs"
        ws.append(NOC_COLUMNS)
        self._create_sheet("scrape_log")
        self._create_sheet("airports")
        self._seed_airports()
        self._save()

    def _create_sheet(self, name: str) -> None:
        cols = {"nocs": NOC_COLUMNS, "scrape_log": LOG_COLUMNS, "airports": AIRPORT_COLUMNS}[name]
        ws = self.wb.create_sheet(name)
        ws.append(cols)

    def _seed_airports(self) -> None:
        ws = self.wb["airports"]
        # Minimal MMR seed — matches db/seed_airports.sql but lighter for Excel
        rows = [
            (15,  "VABB", "Chhatrapati Shivaji Maharaj Intl (Santa Cruz)", 19.0916944, 72.8654722, 11.28),
            (16,  "VAJJ", "Juhu Aerodrome",                                 19.0967,    72.8347,    4.5),
            (140, "VANM", "Navi Mumbai International Airport",              19.0567,    73.0250,    3.0),
        ]
        for row in rows:
            ws.append(row)

    def _save(self) -> None:
        # Atomic-ish: write to .tmp, rename. openpyxl doesn't support streaming
        # so we have to rewrite the whole workbook each save; for ~20k rows
        # that's ~1-2s per save which is fine at 90 windows / hour.
        tmp = self.path.with_suffix(self.path.suffix + ".tmp")
        self.wb.save(tmp)
        os.replace(tmp, self.path)

    def _airport_code(self, aai_id: int) -> str:
        return AAI_ID_TO_CODE.get(aai_id, f"AAI_{aai_id}")

    def get_airport_id(self, aai_id: int) -> int:
        # For Excel, the "airport id" is just the AAI ID — no separate FK.
        return aai_id

    def _load_existing_noc_ids(self) -> set[str]:
        if self._existing_noc_ids is not None:
            return self._existing_noc_ids
        ws = self.wb["nocs"]
        ids: set[str] = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                ids.add(str(row[0]))
        self._existing_noc_ids = ids
        return ids

    def upsert_nocs(self, airport_id: int, parsed: list["ParsedNOC"]) -> tuple[int, int]:
        if not parsed:
            return 0, 0
        ws = self.wb["nocs"]
        existing = self._load_existing_noc_ids()
        airport_code = self._airport_code(airport_id)

        added = 0
        skipped = 0
        for p in parsed:
            if p.noc_id in existing:
                skipped += 1
                continue
            row = noc_to_row(p, airport_code)
            ws.append([row[col] for col in NOC_COLUMNS])
            existing.add(p.noc_id)
            added += 1
        self._save()
        return added, skipped

    def log_run_start(
        self,
        scraper_name: str,
        window: "ScrapeWindow",
        airport_id: int,
        tracking_id: Optional[str],
        trigger_source: str,
    ) -> str:
        run_id = str(uuid.uuid4())
        ws = self.wb["scrape_log"]
        ws.append([
            run_id,
            tracking_id or "",
            scraper_name,
            trigger_source,
            self._airport_code(airport_id),
            window.structure_type,
            window.from_date.isoformat(),
            window.to_date.isoformat(),
            _now_iso(),  # started_at
            "",          # completed_at
            "",          # duration_seconds
            0, 0, 0,
            "running",
            "",
        ])
        self._save()
        return run_id

    def log_run_complete(
        self,
        run_id: str,
        added: int,
        updated: int,
        errors: int,
        duration_ms: int,
        status: str,
        notes: str = "",
    ) -> None:
        ws = self.wb["scrape_log"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == run_id:
                row[9].value  = _now_iso()                    # completed_at
                row[10].value = round(duration_ms / 1000.0, 2)  # duration_seconds
                row[11].value = added                          # records_added
                row[12].value = updated                        # records_updated
                row[13].value = errors                         # errors
                row[14].value = status
                row[15].value = notes
                break
        self._save()

    def close(self) -> None:
        self._save()
        logger.info("Workbook saved: %s", self.path)


# ===========================================================================
# CSV writer (simplest possible)
# ===========================================================================

class CsvWriter:
    """Single-file CSV output. No scrape_log; no airports sheet."""

    def __init__(self, path: Path) -> None:
        self.path = Path(path).expanduser().resolve()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._existing_noc_ids: set[str] = set()
        self._initialize_or_load()

    def _initialize_or_load(self) -> None:
        if self.path.exists():
            with self.path.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get("noc_id"):
                        self._existing_noc_ids.add(row["noc_id"])
            logger.info("Loaded %d existing rows from %s", len(self._existing_noc_ids), self.path)
        else:
            with self.path.open("w", encoding="utf-8", newline="") as f:
                csv.DictWriter(f, fieldnames=NOC_COLUMNS).writeheader()
            logger.info("Created new CSV: %s", self.path)

    def get_airport_id(self, aai_id: int) -> int:
        return aai_id

    def log_run_start(self, *args, **kwargs) -> str:
        # CSV mode skips run logging; return a synthetic ID for the orchestrator.
        return str(uuid.uuid4())

    def log_run_complete(self, *args, **kwargs) -> None:
        pass  # nothing to update

    def upsert_nocs(self, airport_id: int, parsed: list["ParsedNOC"]) -> tuple[int, int]:
        if not parsed:
            return 0, 0
        airport_code = AAI_ID_TO_CODE.get(airport_id, f"AAI_{airport_id}")
        added = 0
        skipped = 0
        with self.path.open("a", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=NOC_COLUMNS)
            for p in parsed:
                if p.noc_id in self._existing_noc_ids:
                    skipped += 1
                    continue
                writer.writerow(noc_to_row(p, airport_code))
                self._existing_noc_ids.add(p.noc_id)
                added += 1
        return added, skipped

    def close(self) -> None:
        logger.info("CSV closed: %s (%d rows)", self.path, len(self._existing_noc_ids))


# ===========================================================================
# Supabase writer (production target)
# ===========================================================================

class SupabaseWriter:
    """Wraps the existing supabase-py logic in the Writer protocol."""

    def __init__(self) -> None:
        try:
            from supabase import create_client
        except ImportError as e:
            raise RuntimeError("supabase-py not installed. pip install supabase") from e
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        if not url or not key:
            raise RuntimeError(
                "Set SUPABASE_URL and SUPABASE_KEY (service role) env vars, "
                "or use --output excel:path / csv:path."
            )
        self.sb = create_client(url, key)
        self._airport_cache: dict[int, int] = {}

    def get_airport_id(self, aai_id: int) -> int:
        if aai_id in self._airport_cache:
            return self._airport_cache[aai_id]
        resp = self.sb.table("airports").select("id").eq("aai_id", aai_id).single().execute()
        if not resp.data:
            raise RuntimeError(
                f"Airport with aai_id={aai_id} not in airports table. Run db/seed_airports.sql first."
            )
        db_id = int(resp.data["id"])
        self._airport_cache[aai_id] = db_id
        return db_id

    def log_run_start(
        self,
        scraper_name: str,
        window: "ScrapeWindow",
        airport_id: int,
        tracking_id: Optional[str],
        trigger_source: str,
    ) -> str:
        row = {
            "scraper":        scraper_name,
            "airport_id":     airport_id,
            "structure_type": window.structure_type,
            "from_date":      window.from_date.isoformat(),
            "to_date":        window.to_date.isoformat(),
            "status":         "running",
            "trigger_source": trigger_source,
        }
        if tracking_id:
            row["tracking_id"] = tracking_id
        resp = self.sb.table("scrape_log").insert(row).execute()
        return str(resp.data[0]["id"])

    def log_run_complete(
        self,
        run_id: str,
        added: int,
        updated: int,
        errors: int,
        duration_ms: int,
        status: str,
        notes: str = "",
    ) -> None:
        self.sb.table("scrape_log").update({
            "completed_at":    datetime.utcnow().isoformat(),
            "duration_ms":     duration_ms,
            "records_added":   added,
            "records_updated": updated,
            "errors":          errors,
            "status":          status,
            "notes":           notes,
        }).eq("id", run_id).execute()

    def upsert_nocs(self, airport_id: int, parsed: list["ParsedNOC"]) -> tuple[int, int]:
        if not parsed:
            return 0, 0
        rows = []
        for p in parsed:
            rows.append({
                "noc_id":              p.noc_id,
                "sacfa_id":            p.sacfa_id,
                "airport_id":          airport_id,
                "lat":                 p.lat,
                "lon":                 p.lon,
                "site_elevation_m":    p.site_elevation_m or 0,
                "permissible_top_m":   p.permissible_top_m,
                "permissible_top_raw": p.permissible_top_raw,
                "is_restricted":       p.is_restricted,
                "structure_type":      p.structure_type,
                "status":              p.status,
                "issue_date":          p.issue_date.isoformat() if p.issue_date else None,
                "issue_date_known":    p.issue_date is not None,
                "pdf_url":             p.pdf_url,
                "owner_name":          p.owner_name,
                "site_address":        p.site_address,
                "raw_payload":         p.raw_payload,
            })
        self.sb.table("noc_issued").upsert(rows, on_conflict="noc_id").execute()
        return len(rows), 0

    def close(self) -> None:
        pass


# ---------------------------------------------------------------------------
# Factory
# ---------------------------------------------------------------------------

def make_writer(spec: str) -> OutputWriter:
    """
    spec examples:
        "excel:./data/skygauge.xlsx"
        "csv:./data/skygauge.csv"
        "supabase"
        "auto"  — picks Supabase if env vars set, else Excel default
    """
    if spec == "auto":
        if os.environ.get("SUPABASE_URL") and (
            os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        ):
            spec = "supabase"
        else:
            spec = "excel:./data/skygauge_nocs.xlsx"

    if spec == "supabase":
        return SupabaseWriter()

    if ":" in spec:
        mode, path = spec.split(":", 1)
        if mode == "excel":
            return ExcelWriter(Path(path))
        if mode == "csv":
            return CsvWriter(Path(path))

    raise ValueError(
        f"Unrecognised --output: {spec!r}. "
        f"Use 'excel:path.xlsx', 'csv:path.csv', 'supabase', or 'auto'."
    )
