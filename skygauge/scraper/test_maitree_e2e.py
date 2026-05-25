"""End-to-end smoke test for maitree_scraper.run().

Monkey-patches `fetch_heightnoc` so no network is needed, then exercises:
  fetch → extract → bucket-by-airport → write to Excel → write appeals CSV

Run from the skygauge/ directory:
    python scraper/test_maitree_e2e.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import maitree_scraper  # noqa: E402
from maitree_scraper import FetchResult, run  # noqa: E402

OUT_XLSX = Path("/tmp/skygauge_maitree_e2e.xlsx")
OUT_CSV = Path("/tmp/skygauge_maitree_e2e_appeals.csv")


SYNTHETIC_HTML = """<html><body>
<script>
    //var locations = [{"PermanentNOCID":"STALE/STUB","Lat":"0","Long":"0"}];
    var locations = [
        {"PermanentNOCID":"JUHU/WEST/B/020616/120684","IssueDate":"29/03/16",
         "Lat":"19.14415833","Long":"72.83776944","SiteElevation":"10.97",
         "PermissibleTopElevation":"100.58 M (Restricted)",
         "Link":"https://nocas2.aai.aero/nocas/NOC_Letters/JUHU_WEST_B_020616_120684.pdf"},
        {"PermanentNOCID":"SNCR/WEST/B/020116/120467","IssueDate":"01/04/16",
         "Lat":"19.10995","Long":"72.86789167","SiteElevation":"16.32",
         "PermissibleTopElevation":"51.32M",
         "Link":"https://nocas2.aai.aero/nocas/NOC_Letters/SNCR_WEST_B_020116_120467.pdf"},
        {"PermanentNOCID":"SNCR/WEST/M/030516/121005","IssueDate":"15/05/16",
         "Lat":"19.07217222","Long":"72.87987222","SiteElevation":"6.5",
         "PermissibleTopElevation":"41.8M",
         "Link":"https://nocas2.aai.aero/nocas/NOC_Letters/SNCR_WEST_M_030516_121005.pdf"},
        {"PermanentNOCID":"NMIA/EAST/B/070124/000001","IssueDate":"07/01/24",
         "Lat":"19.05670000","Long":"73.02500000","SiteElevation":"3.0",
         "PermissibleTopElevation":"45.0 M",
         "Link":"https://nocas2.aai.aero/nocas/NOC_Letters/NMIA_EAST_B_070124_000001.pdf"},
        {"PermanentNOCID":"","IssueDate":"","Lat":"","Long":"","SiteElevation":"",
         "PermissibleTopElevation":"","Link":""}
    ];
    var appealCases = [
        {"NOC_ID":"SNCR/WEST/B/050624/881234","Date":"10/12/25","Item_No":"22",
         "Lat":"19.04059167","Long":"72.87055278",
         "Site_Elevation_AMSL_in_M":"4.1",
         "Permissible_Top_Elevation_Approved_by_Committee":"140",
         "Link":"https://nocas2.aai.aero/Nocas/AppealProceeding/foo.pdf"},
        {"NOC_ID":"JUHU/WEST/B/060724/901111","Date":"15/08/25","Item_No":"7",
         "Lat":"19.10000000","Long":"72.85000000",
         "Site_Elevation_AMSL_in_M":"5.5",
         "Permissible_Top_Elevation_Approved_by_Committee":"75",
         "Link":"https://nocas2.aai.aero/Nocas/AppealProceeding/bar.pdf"}
    ];
</script>
</body></html>"""


def _fake_fetch(*args, **kwargs) -> FetchResult:
    return FetchResult(
        html=SYNTHETIC_HTML,
        elapsed_seconds=0.01,
        size_bytes=len(SYNTHETIC_HTML),
    )


def main() -> int:
    for p in (OUT_XLSX, OUT_CSV):
        if p.exists():
            p.unlink()

    # Patch network call
    maitree_scraper.fetch_heightnoc = _fake_fetch

    summary = run(
        output=f"excel:{OUT_XLSX}",
        appeals_csv_path=OUT_CSV,
        dry_run=False,
    )

    print(summary)

    # Assertions
    assert summary.locations_total == 5
    assert summary.locations_parsed == 4, "Expected 4 (1 empty placeholder skipped)"
    assert summary.locations_unknown_airport == 0
    assert summary.locations_parse_failures == 1, "The empty record should count as a parse failure"
    assert summary.nocs_added == 4
    assert summary.nocs_skipped == 0  # fresh workbook
    assert summary.appeals_total == 2
    assert summary.appeals_written == 2
    print("[ok] run summary matches expectations")

    # Verify workbook shape
    import openpyxl
    wb = openpyxl.load_workbook(OUT_XLSX)
    assert "nocs" in wb.sheetnames
    assert "scrape_log" in wb.sheetnames
    assert "airports" in wb.sheetnames
    nocs = wb["nocs"]
    assert nocs.max_row == 5, f"Expected 1 header + 4 records, got {nocs.max_row}"
    headers = [c.value for c in nocs[1]]
    rows_by_id: dict[str, dict] = {}
    for row in nocs.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows_by_id[str(record["noc_id"])] = record

    # Spot-check the JUHU record's elevation/top/pdf
    juhu = rows_by_id["JUHU/WEST/B/020616/120684"]
    assert juhu["site_elevation_m"] == 10.97, f"JUHU elev {juhu['site_elevation_m']!r}"
    assert juhu["permissible_top_m"] == 100.58, f"JUHU top {juhu['permissible_top_m']!r}"
    assert juhu["is_restricted"] is True
    assert juhu["airport_code"] == "VAJJ"
    assert "JUHU_WEST_B_020616_120684.pdf" in juhu["pdf_url"]

    # And the NMIA record (different airport bucket)
    nmia = rows_by_id["NMIA/EAST/B/070124/000001"]
    assert nmia["airport_code"] == "VANM"
    assert nmia["site_elevation_m"] == 3.0
    print(f"[ok] workbook has {nocs.max_row - 1} NOC rows with elev/top/pdf populated; "
          f"airport_codes routed correctly")

    # Verify scrape_log has one row per airport that received records
    sl = wb["scrape_log"]
    assert sl.max_row == 4, f"Expected 1 header + 3 airport runs, got {sl.max_row}"
    print(f"[ok] scrape_log has {sl.max_row - 1} run rows (one per airport)")

    # Verify appeals CSV
    import csv
    with OUT_CSV.open() as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 2
    assert float(rows[0]["site_elevation_m"]) == 4.1
    assert float(rows[0]["permissible_top_committee_m"]) == 140.0
    print(f"[ok] appeals CSV has {len(rows)} rows with elevation populated")

    # Idempotency check
    summary2 = run(
        output=f"excel:{OUT_XLSX}",
        appeals_csv_path=OUT_CSV,
        dry_run=False,
    )
    assert summary2.nocs_added == 0, f"Second run should add 0, got {summary2.nocs_added}"
    assert summary2.nocs_skipped == 4, f"Second run should skip 4 as dupes, got {summary2.nocs_skipped}"
    print("[ok] second run is idempotent (4 dupes skipped, 0 added)")

    print("\nE2E test passed. Workbook at:", OUT_XLSX)
    return 0


if __name__ == "__main__":
    sys.exit(main())
