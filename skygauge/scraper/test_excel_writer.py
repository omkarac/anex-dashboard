"""Smoke test for the Excel writer with synthetic AAI-shaped records.

Run from the skygauge/ directory:
    python scraper/test_excel_writer.py

Output: /tmp/skygauge_test.xlsx (3 sheets: nocs, scrape_log, airports)
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from nocas_scraper import ScrapeWindow, parse_noc_record  # noqa: E402
from writers import ExcelWriter  # noqa: E402


# Synthetic records in the exact shape we observed on Project Maitree
# (which is what they got from the same AAI endpoint we're hitting).
SYNTHETIC_RECORDS: list[dict] = [
    {
        "PermanentNOCID": "SNCR_WEST_B_011625_1494190",
        "IssueDate": "12/03/26",
        "Lat": "19.061175",
        "Long": "72.88020278",
        "SiteElevation": "4.32",
        "PermissibleTopElevation": "49.32 M",
        "Link": "https://nocas2.aai.aero/nocas/NOC_Letters/SNCR_WEST_B_011625_1494190.pdf",
    },
    {
        "PermanentNOCID": "JUHU_WEST_B_020616_120684",
        "IssueDate": "29/03/16",
        "Lat": "19.14415833",
        "Long": "72.83776944",
        "SiteElevation": "10.97",
        "PermissibleTopElevation": "100.58 M (Restricted)",
        "Link": "https://nocas2.aai.aero/nocas/NOC_Letters/JUHU_WEST_B_020616_120684.pdf",
    },
    {
        "PermanentNOCID": "BKC_B_2024_0042",
        "IssueDate": "15/08/24",
        "Lat": "19.0660",
        "Long": "72.8680",
        "SiteElevation": "8.5",
        "PermissibleTopElevation": "56.0 M",
        "Link": "https://nocas2.aai.aero/nocas/NOC_Letters/BKC_B_2024_0042.pdf",
    },
]


def main() -> int:
    out_path = Path("/tmp/skygauge_test.xlsx")
    if out_path.exists():
        out_path.unlink()

    writer = ExcelWriter(out_path)

    # Simulated window: Santa Cruz, Buildings, 2024 full year
    window = ScrapeWindow(
        airport_aai_id=15,
        structure_type="B",
        from_date=date(2024, 1, 1),
        to_date=date(2024, 12, 31),
    )

    airport_id = writer.get_airport_id(15)
    run_id = writer.log_run_start(
        scraper_name="nocas",
        window=window,
        airport_id=airport_id,
        tracking_id=None,
        trigger_source="backfill",
    )
    print(f"Started run {run_id}")

    parsed = [parse_noc_record(r, "B") for r in SYNTHETIC_RECORDS]
    parsed = [p for p in parsed if p]
    print(f"Parsed {len(parsed)} synthetic records")

    added, skipped = writer.upsert_nocs(airport_id, parsed)
    print(f"upsert_nocs → added={added}, skipped={skipped}")

    writer.log_run_complete(run_id, added, skipped, 0, 1234, "success", "smoke test")

    # Re-run to verify idempotency
    added2, skipped2 = writer.upsert_nocs(airport_id, parsed)
    print(f"Re-run idempotency check → added={added2} (expect 0), skipped={skipped2} (expect {len(parsed)})")
    assert added2 == 0, "Expected idempotent: 0 new rows on second insert"
    assert skipped2 == len(parsed)

    writer.close()
    print(f"\n✓ Wrote {out_path}")
    print(f"  Size: {out_path.stat().st_size} bytes")

    # Read back and verify
    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    print(f"  Sheets: {wb.sheetnames}")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"  {sheet}: {ws.max_row} rows × {ws.max_column} cols")

    print("\nFirst NOC row:")
    nocs_ws = wb["nocs"]
    headers = [c.value for c in nocs_ws[1]]
    first_data = [c.value for c in nocs_ws[2]]
    for h, v in zip(headers, first_data):
        print(f"  {h:24s} = {v}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
